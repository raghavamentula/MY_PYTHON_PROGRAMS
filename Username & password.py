import openpyxl
wb= openpyxl.load_workbook("D:/Rahgava Python/Games/user & pass.xlsx")
sh = wb['Sheet1']
true = True
Us= str(input('Hi! creat your username').lower())
row = sh.max_row
for i in range (2,row+1):
    lo = 'A'+str(i)
    user = sh[lo]
    if user.value == Us:
        print('sorry alredy exists try another username')
        true = False
        break
if true == True:
    Password = input('Okay write password secretly!!!!!!!!!!!')
    sh.cell(row+1,2,value=Password)
    sh.cell(row+1,1,value=Us)
    wb.save('D:/Rahgava Python/Games/user & pass.xlsx')
